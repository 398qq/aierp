"""Generic import/export endpoints for Excel-based batch operations."""

import csv
import io
import datetime

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl

from app.api.deps import get_current_user
from app.database import get_db
from app.models.customer import Customer
from app.models.finance import Contract
from app.models.product import Brand, Product, Supplier
from app.models.sales import Quotation, SalesOrder
from app.models.transaction import PurchaseOrder
from app.schemas.common import fail, ok

router = APIRouter(tags=["import-export"])

EXPORT_ENTITIES = ["customers", "products", "suppliers", "brands", "purchase_orders", "sales_orders", "quotations", "contracts"]


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
@router.get("/export/{entity}")
async def export_entity(
    entity: str,
    format: str = Query("csv"),
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    if entity not in EXPORT_ENTITIES:
        return fail(f"不支持的实体类型: {entity}")

    if entity == "customers":
        result = await db.execute(
            select(Customer).where(Customer.deleted_at.is_(None)).order_by(Customer.id)
        )
        rows = result.scalars().all()
        headers = ["id", "name", "code", "phone", "email", "industry", "level", "source", "notes"]
        data = [[r.id, r.name, r.code, r.phone, r.email, r.industry, r.level, r.source, r.notes] for r in rows]
    elif entity == "products":
        result = await db.execute(
            select(Product).where(Product.deleted_at.is_(None)).order_by(Product.id)
        )
        rows = result.scalars().all()
        headers = ["id", "name", "sku", "category", "brand_id", "cost_price", "selling_price", "unit"]
        data = [[r.id, r.name, r.sku, r.category, r.brand_id, float(r.cost_price),
                 float(r.selling_price), r.unit] for r in rows]
    elif entity == "suppliers":
        result = await db.execute(
            select(Supplier).where(Supplier.deleted_at.is_(None)).order_by(Supplier.id)
        )
        rows = result.scalars().all()
        headers = ["id", "name", "code", "contact", "phone", "email", "address", "level"]
        data = [[r.id, r.name, r.code, r.contact, r.phone, r.email, r.address, r.level] for r in rows]
    elif entity == "brands":
        result = await db.execute(
            select(Brand).where(Brand.deleted_at.is_(None)).order_by(Brand.id)
        )
        rows = result.scalars().all()
        headers = ["id", "name", "description"]
        data = [[r.id, r.name, r.description] for r in rows]
    elif entity == "purchase_orders":
        result = await db.execute(
            select(PurchaseOrder).where(PurchaseOrder.deleted_at.is_(None)).order_by(PurchaseOrder.id)
        )
        rows = result.scalars().all()
        headers = ["id", "order_no", "supplier_id", "total_amount", "status", "created_at"]
        data = [[r.id, r.order_no, r.supplier_id, float(r.total_amount), r.status, str(r.created_at)] for r in rows]
    elif entity == "sales_orders":
        result = await db.execute(
            select(SalesOrder).where(SalesOrder.deleted_at.is_(None)).order_by(SalesOrder.id)
        )
        rows = result.scalars().all()
        headers = ["id", "order_no", "customer_id", "total_amount", "status", "created_at"]
        data = [[r.id, r.order_no, r.customer_id, float(r.total_amount), r.status, str(r.created_at)] for r in rows]
    elif entity == "contracts":
        result = await db.execute(
            select(Contract).where(Contract.deleted_at.is_(None)).order_by(Contract.id)
        )
        rows = result.scalars().all()
        headers = ["id", "contract_no", "title", "customer_id", "sales_order_id", "amount", "signed_date", "expire_date", "status", "notes"]
        data = [[r.id, r.contract_no, r.title, r.customer_id, r.sales_order_id, float(r.amount),
                 str(r.signed_date) if r.signed_date else "", str(r.expire_date) if r.expire_date else "",
                 r.status, r.notes or ""] for r in rows]
    else:
        result = await db.execute(
            select(Quotation).where(Quotation.deleted_at.is_(None)).order_by(Quotation.id)
        )
        rows = result.scalars().all()
        headers = ["id", "title", "customer_id", "total_amount", "status", "created_at"]
        data = [[r.id, r.title, r.customer_id, float(r.total_amount), r.status, str(r.created_at)] for r in rows]

    if format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in data:
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={entity}_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"},
        )
    else:
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(headers)
        w.writerows(data)
        out.seek(0)
        return StreamingResponse(
            io.BytesIO(out.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={entity}_{datetime.datetime.now().strftime('%Y%m%d')}.csv"},
        )


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
IMPORT_ENTITIES = ["customers", "products", "suppliers", "contracts"]


@router.post("/import/{entity}")
async def import_entity(
    entity: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    if entity not in IMPORT_ENTITIES:
        return fail(f"不支持的实体类型: {entity}")

    content = await file.read()

    if file.filename and file.filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
    else:
        reader = csv.reader(io.StringIO(content.decode("utf-8-sig")))
        rows_iter = reader

    headers = next(rows_iter, None)
    if not headers:
        return fail("文件为空或缺少标题行")

    created = 0
    errors = []

    for i, row in enumerate(rows_iter):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            row_dict = {str(headers[j]): str(row[j]) if j < len(row) else "" for j in range(len(headers))}

            if entity == "customers":
                c = Customer(
                    name=row_dict.get("name", row_dict.get("名称", "")),
                    phone=row_dict.get("phone", row_dict.get("电话", "")),
                    email=row_dict.get("email", ""),
                    industry=row_dict.get("industry", row_dict.get("行业", "")),
                    level=row_dict.get("level", "C"),
                    source=row_dict.get("source", ""),
                )
                db.add(c)
            elif entity == "products":
                p = Product(
                    name=row_dict.get("name", row_dict.get("名称", "")),
                    sku=row_dict.get("sku", ""),
                    category=row_dict.get("category", ""),
                    cost_price=float(row_dict.get("cost_price", 0)),
                    selling_price=float(row_dict.get("selling_price", 0)),
                    unit=row_dict.get("unit", "pcs"),
                )
                db.add(p)
            elif entity == "suppliers":
                s = Supplier(
                    name=row_dict.get("name", row_dict.get("名称", "")),
                    contact=row_dict.get("contact", ""),
                    phone=row_dict.get("phone", ""),
                    email=row_dict.get("email", ""),
                    address=row_dict.get("address", ""),
                    level=row_dict.get("level", "B"),
                )
                db.add(s)
            elif entity == "contracts":
                signed_str = row_dict.get("signed_date", row_dict.get("签署日期", ""))
                expire_str = row_dict.get("expire_date", row_dict.get("到期日期", ""))
                ct = Contract(
                    contract_no=row_dict.get("contract_no", row_dict.get("合同号", "")),
                    title=row_dict.get("title", row_dict.get("标题", "")),
                    customer_id=int(row_dict.get("customer_id", 0)),
                    sales_order_id=int(row_dict.get("sales_order_id")) if row_dict.get("sales_order_id") else None,
                    amount=float(row_dict.get("amount", row_dict.get("金额", 0))),
                    signed_date=datetime.date.fromisoformat(signed_str) if signed_str else None,
                    expire_date=datetime.date.fromisoformat(expire_str) if expire_str else None,
                    status=row_dict.get("status", row_dict.get("状态", "draft")),
                    notes=row_dict.get("notes", row_dict.get("备注", "")),
                )
                db.add(ct)
            created += 1
        except Exception as e:
            errors.append(f"行 {i + 2}: {str(e)}")

    await db.commit()
    return ok({
        "created": created, "errors": errors, "total_rows": i + 1 if 'i' in dir() else 0,
    })
