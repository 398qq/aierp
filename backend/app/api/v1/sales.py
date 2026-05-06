"""Sales management API — opportunities, quotations, sales orders, delivery notes."""

import io
from datetime import datetime, timezone

import openpyxl
from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.database import get_db
from app.models.sales import (
    DeliveryNote, DeliveryNoteItem, Opportunity, Quotation, QuotationItem, SalesOrder, SalesOrderItem,
)
from app.schemas.common import fail, ok
from app.schemas.sales import (
    BatchDeleteRequest,
    DeliveryNoteCreate, DeliveryNoteItemCreate, DeliveryNoteItemUpdate, DeliveryNoteUpdate,
    OpportunityBatchUpdate, OpportunityCreate, OpportunityUpdate,
    QuotationCreate, QuotationItemCreate, QuotationItemUpdate, QuotationUpdate,
    SalesOrderCreate, SalesOrderItemCreate, SalesOrderItemUpdate, SalesOrderUpdate,
)

router = APIRouter(tags=["sales"])


def _parse_date(value: str | None) -> datetime | None:
    if value:
        return datetime.fromisoformat(value)
    return None


def _serialize_dt(dt: datetime | None) -> str | None:
    return str(dt) if dt else None


async def _generate_doc_no(db: AsyncSession, prefix: str, model) -> str:
    """Generate document number like QO-20260506-001 by auto-detecting _no column."""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    pattern = f"{prefix}-{today}-%"
    # auto-detect column ending with '_no'
    no_col_name = next(
        c.name for c in model.__table__.columns if c.name.endswith("_no")
    )
    col = getattr(model, no_col_name)
    result = await db.execute(
        select(func.count(model.id)).where(col.like(pattern), model.deleted_at.is_(None))
    )
    seq = (result.scalar() or 0) + 1
    return f"{prefix}-{today}-{seq:03d}"


async def _next_seq(db: AsyncSession, prefix: str, model, no_col: str) -> str:
    """Generate document number by counting today's records."""
    today = datetime.now(timezone.utc).strftime("%Y%m%d")
    pattern = f"{prefix}-{today}-%"
    col = getattr(model, no_col)
    result = await db.execute(
        select(func.count(model.id)).where(col.like(pattern), model.deleted_at.is_(None))
    )
    cnt = (result.scalar() or 0) + 1
    return f"{prefix}-{today}-{cnt:03d}"


# --- Opportunities ---

opp_router = APIRouter(prefix="/opportunities")


def _opp_row(o: Opportunity) -> dict:
    return {
        "id": o.id, "customer_id": o.customer_id, "name": o.name,
        "amount": float(o.amount), "stage": o.stage, "probability": o.probability,
        "expected_close_date": _serialize_dt(o.expected_close_date),
        "actual_close_date": _serialize_dt(o.actual_close_date),
        "notes": o.notes, "created_at": str(o.created_at),
    }


@opp_router.get("")
async def list_opportunities(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    customer_id: int | None = None,
    stage: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    base = select(Opportunity).where(Opportunity.deleted_at.is_(None))
    count_base = select(func.count(Opportunity.id)).where(Opportunity.deleted_at.is_(None))

    if customer_id:
        base = base.where(Opportunity.customer_id == customer_id)
        count_base = count_base.where(Opportunity.customer_id == customer_id)
    if stage:
        base = base.where(Opportunity.stage == stage)
        count_base = count_base.where(Opportunity.stage == stage)

    total = (await db.execute(count_base)).scalar() or 0
    rows = (await db.execute(
        base.order_by(Opportunity.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )).scalars().all()

    return ok({
        "list": [_opp_row(o) for o in rows],
        "total": total, "page": page, "page_size": page_size,
    })


@opp_router.get("/{opp_id}")
async def get_opportunity(opp_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(Opportunity).where(Opportunity.id == opp_id, Opportunity.deleted_at.is_(None))
    )
    opp = result.scalar_one_or_none()
    if opp is None:
        return fail("Opportunity not found", 404)
    return ok(_opp_row(opp))


@opp_router.post("", status_code=201)
async def create_opportunity(body: OpportunityCreate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    data = body.model_dump()
    for date_field in ("expected_close_date", "actual_close_date"):
        if data.get(date_field):
            data[date_field] = _parse_date(data[date_field])
    opp = Opportunity(**data)
    db.add(opp)
    await db.flush()
    return ok({"id": opp.id, "name": opp.name})


@opp_router.put("/{opp_id}")
async def update_opportunity(opp_id: int, body: OpportunityUpdate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(Opportunity).where(Opportunity.id == opp_id, Opportunity.deleted_at.is_(None))
    )
    opp = result.scalar_one_or_none()
    if opp is None:
        return fail("Opportunity not found", 404)
    data = body.model_dump(exclude_unset=True)
    for date_field in ("expected_close_date", "actual_close_date"):
        if data.get(date_field):
            data[date_field] = _parse_date(data[date_field])
    for key, val in data.items():
        setattr(opp, key, val)
    await db.flush()
    return ok({"id": opp.id})


@opp_router.delete("/{opp_id}")
async def delete_opportunity(opp_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(Opportunity).where(Opportunity.id == opp_id, Opportunity.deleted_at.is_(None))
    )
    opp = result.scalar_one_or_none()
    if opp is None:
        return fail("Opportunity not found", 404)
    opp.deleted_at = datetime.now(timezone.utc)
    await db.flush()
    return ok(msg="deleted")


# --- Quotations ---

quo_router = APIRouter(prefix="/quotations")


def _quo_row(q: Quotation) -> dict:
    return {
        "id": q.id, "quotation_no": q.quotation_no, "customer_id": q.customer_id,
        "status": q.status, "total_amount": float(q.total_amount),
        "valid_until": _serialize_dt(q.valid_until),
        "notes": q.notes, "created_at": str(q.created_at),
    }


@quo_router.get("")
async def list_quotations(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    customer_id: int | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    base = select(Quotation).where(Quotation.deleted_at.is_(None))
    count_base = select(func.count(Quotation.id)).where(Quotation.deleted_at.is_(None))

    if customer_id:
        base = base.where(Quotation.customer_id == customer_id)
        count_base = count_base.where(Quotation.customer_id == customer_id)
    if status:
        base = base.where(Quotation.status == status)
        count_base = count_base.where(Quotation.status == status)

    total = (await db.execute(count_base)).scalar() or 0
    rows = (await db.execute(
        base.order_by(Quotation.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )).scalars().all()

    return ok({
        "list": [_quo_row(q) for q in rows],
        "total": total, "page": page, "page_size": page_size,
    })


@quo_router.get("/{quo_id}")
async def get_quotation(quo_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(Quotation).where(Quotation.id == quo_id, Quotation.deleted_at.is_(None))
    )
    quo = result.scalar_one_or_none()
    if quo is None:
        return fail("Quotation not found", 404)
    data = _quo_row(quo)
    items_result = await db.execute(
        select(QuotationItem).where(QuotationItem.quotation_id == quo_id, QuotationItem.deleted_at.is_(None))
    )
    data["items"] = [{
        "id": i.id, "quotation_id": i.quotation_id, "product_id": i.product_id,
        "quantity": i.quantity, "unit_price": float(i.unit_price), "amount": float(i.amount),
    } for i in items_result.scalars().all()]
    return ok(data)


@quo_router.post("", status_code=201)
async def create_quotation(body: QuotationCreate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    data = body.model_dump()
    if not data.get("quotation_no"):
        data["quotation_no"] = await _next_seq(db, "QO", Quotation, "quotation_no")
    if data.get("valid_until"):
        data["valid_until"] = _parse_date(data["valid_until"])
    quo = Quotation(**data)
    db.add(quo)
    await db.flush()
    return ok({"id": quo.id, "quotation_no": quo.quotation_no})


@quo_router.put("/{quo_id}")
async def update_quotation(quo_id: int, body: QuotationUpdate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(Quotation).where(Quotation.id == quo_id, Quotation.deleted_at.is_(None))
    )
    quo = result.scalar_one_or_none()
    if quo is None:
        return fail("Quotation not found", 404)
    data = body.model_dump(exclude_unset=True)
    if data.get("valid_until"):
        data["valid_until"] = _parse_date(data["valid_until"])
    for key, val in data.items():
        setattr(quo, key, val)
    await db.flush()
    return ok({"id": quo.id})


@quo_router.delete("/{quo_id}")
async def delete_quotation(quo_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(Quotation).where(Quotation.id == quo_id, Quotation.deleted_at.is_(None))
    )
    quo = result.scalar_one_or_none()
    if quo is None:
        return fail("Quotation not found", 404)
    quo.deleted_at = datetime.now(timezone.utc)
    await db.flush()
    return ok(msg="deleted")


# --- Quotation Items ---

@quo_router.get("/{quo_id}/items")
async def list_quotation_items(quo_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    rows = (await db.execute(
        select(QuotationItem).where(
            QuotationItem.quotation_id == quo_id, QuotationItem.deleted_at.is_(None)
        )
    )).scalars().all()
    return ok([{
        "id": i.id, "quotation_id": i.quotation_id, "product_id": i.product_id,
        "quantity": i.quantity, "unit_price": float(i.unit_price), "amount": float(i.amount),
    } for i in rows])


@quo_router.post("/{quo_id}/items", status_code=201)
async def create_quotation_item(quo_id: int, body: QuotationItemCreate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    item = QuotationItem(quotation_id=quo_id, **body.model_dump())
    db.add(item)
    await db.flush()
    return ok({"id": item.id, "product_id": item.product_id})


@quo_router.put("/{quo_id}/items/{item_id}")
async def update_quotation_item(quo_id: int, item_id: int, body: QuotationItemUpdate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(QuotationItem).where(
            QuotationItem.id == item_id, QuotationItem.quotation_id == quo_id,
            QuotationItem.deleted_at.is_(None),
        )
    )
    item = result.scalar_one_or_none()
    if item is None:
        return fail("Quotation item not found", 404)
    for key, val in body.model_dump(exclude_unset=True).items():
        setattr(item, key, val)
    await db.flush()
    return ok({"id": item.id})


@quo_router.delete("/{quo_id}/items/{item_id}")
async def delete_quotation_item(quo_id: int, item_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(QuotationItem).where(
            QuotationItem.id == item_id, QuotationItem.quotation_id == quo_id,
            QuotationItem.deleted_at.is_(None),
        )
    )
    item = result.scalar_one_or_none()
    if item is None:
        return fail("Quotation item not found", 404)
    item.deleted_at = datetime.now(timezone.utc)
    await db.flush()
    return ok(msg="deleted")


# --- Sales Orders ---

so_router = APIRouter(prefix="/sales-orders")


def _so_row(s: SalesOrder) -> dict:
    return {
        "id": s.id, "order_no": s.order_no, "customer_id": s.customer_id,
        "status": s.status, "total_amount": float(s.total_amount),
        "delivery_date": _serialize_dt(s.delivery_date),
        "notes": s.notes, "created_at": str(s.created_at),
    }


@so_router.get("")
async def list_sales_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    customer_id: int | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    base = select(SalesOrder).where(SalesOrder.deleted_at.is_(None))
    count_base = select(func.count(SalesOrder.id)).where(SalesOrder.deleted_at.is_(None))

    if customer_id:
        base = base.where(SalesOrder.customer_id == customer_id)
        count_base = count_base.where(SalesOrder.customer_id == customer_id)
    if status:
        base = base.where(SalesOrder.status == status)
        count_base = count_base.where(SalesOrder.status == status)

    total = (await db.execute(count_base)).scalar() or 0
    rows = (await db.execute(
        base.order_by(SalesOrder.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )).scalars().all()

    return ok({
        "list": [_so_row(s) for s in rows],
        "total": total, "page": page, "page_size": page_size,
    })


@so_router.get("/{order_id}")
async def get_sales_order(order_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(SalesOrder).where(SalesOrder.id == order_id, SalesOrder.deleted_at.is_(None))
    )
    order = result.scalar_one_or_none()
    if order is None:
        return fail("Sales order not found", 404)
    data = _so_row(order)
    items_result = await db.execute(
        select(SalesOrderItem).where(SalesOrderItem.order_id == order_id, SalesOrderItem.deleted_at.is_(None))
    )
    data["items"] = [{
        "id": i.id, "order_id": i.order_id, "product_id": i.product_id,
        "quantity": i.quantity, "unit_price": float(i.unit_price), "amount": float(i.amount),
    } for i in items_result.scalars().all()]
    return ok(data)


@so_router.post("", status_code=201)
async def create_sales_order(body: SalesOrderCreate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    data = body.model_dump()
    if not data.get("order_no"):
        data["order_no"] = await _next_seq(db, "SO", SalesOrder, "order_no")
    if data.get("delivery_date"):
        data["delivery_date"] = _parse_date(data["delivery_date"])
    order = SalesOrder(**data)
    db.add(order)
    await db.flush()
    return ok({"id": order.id, "order_no": order.order_no})


@so_router.put("/{order_id}")
async def update_sales_order(order_id: int, body: SalesOrderUpdate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(SalesOrder).where(SalesOrder.id == order_id, SalesOrder.deleted_at.is_(None))
    )
    order = result.scalar_one_or_none()
    if order is None:
        return fail("Sales order not found", 404)
    data = body.model_dump(exclude_unset=True)
    if data.get("delivery_date"):
        data["delivery_date"] = _parse_date(data["delivery_date"])
    for key, val in data.items():
        setattr(order, key, val)
    await db.flush()
    return ok({"id": order.id})


@so_router.delete("/{order_id}")
async def delete_sales_order(order_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(SalesOrder).where(SalesOrder.id == order_id, SalesOrder.deleted_at.is_(None))
    )
    order = result.scalar_one_or_none()
    if order is None:
        return fail("Sales order not found", 404)
    order.deleted_at = datetime.now(timezone.utc)
    await db.flush()
    return ok(msg="deleted")


# --- Sales Order Items ---

@so_router.get("/{order_id}/items")
async def list_sales_order_items(order_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    rows = (await db.execute(
        select(SalesOrderItem).where(
            SalesOrderItem.order_id == order_id, SalesOrderItem.deleted_at.is_(None)
        )
    )).scalars().all()
    return ok([{
        "id": i.id, "order_id": i.order_id, "product_id": i.product_id,
        "quantity": i.quantity, "unit_price": float(i.unit_price), "amount": float(i.amount),
    } for i in rows])


@so_router.post("/{order_id}/items", status_code=201)
async def create_sales_order_item(order_id: int, body: SalesOrderItemCreate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    item = SalesOrderItem(order_id=order_id, **body.model_dump())
    db.add(item)
    await db.flush()
    return ok({"id": item.id, "product_id": item.product_id})


@so_router.put("/{order_id}/items/{item_id}")
async def update_sales_order_item(order_id: int, item_id: int, body: SalesOrderItemUpdate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(SalesOrderItem).where(
            SalesOrderItem.id == item_id, SalesOrderItem.order_id == order_id,
            SalesOrderItem.deleted_at.is_(None),
        )
    )
    item = result.scalar_one_or_none()
    if item is None:
        return fail("Sales order item not found", 404)
    for key, val in body.model_dump(exclude_unset=True).items():
        setattr(item, key, val)
    await db.flush()
    return ok({"id": item.id})


@so_router.delete("/{order_id}/items/{item_id}")
async def delete_sales_order_item(order_id: int, item_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(SalesOrderItem).where(
            SalesOrderItem.id == item_id, SalesOrderItem.order_id == order_id,
            SalesOrderItem.deleted_at.is_(None),
        )
    )
    item = result.scalar_one_or_none()
    if item is None:
        return fail("Sales order item not found", 404)
    item.deleted_at = datetime.now(timezone.utc)
    await db.flush()
    return ok(msg="deleted")


# --- Delivery Notes ---

dn_router = APIRouter(prefix="/delivery-notes")


def _dn_row(d: DeliveryNote) -> dict:
    return {
        "id": d.id, "note_no": d.note_no, "sales_order_id": d.sales_order_id,
        "customer_id": d.customer_id, "status": d.status,
        "delivery_date": _serialize_dt(d.delivery_date),
        "signed_at": _serialize_dt(d.signed_at),
        "notes": d.notes, "created_at": str(d.created_at),
    }


@dn_router.get("")
async def list_delivery_notes(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sales_order_id: int | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    base = select(DeliveryNote).where(DeliveryNote.deleted_at.is_(None))
    count_base = select(func.count(DeliveryNote.id)).where(DeliveryNote.deleted_at.is_(None))

    if sales_order_id:
        base = base.where(DeliveryNote.sales_order_id == sales_order_id)
        count_base = count_base.where(DeliveryNote.sales_order_id == sales_order_id)
    if status:
        base = base.where(DeliveryNote.status == status)
        count_base = count_base.where(DeliveryNote.status == status)

    total = (await db.execute(count_base)).scalar() or 0
    rows = (await db.execute(
        base.order_by(DeliveryNote.id.desc()).offset((page - 1) * page_size).limit(page_size)
    )).scalars().all()

    return ok({
        "list": [_dn_row(d) for d in rows],
        "total": total, "page": page, "page_size": page_size,
    })


@dn_router.get("/{note_id}")
async def get_delivery_note(note_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(DeliveryNote).where(DeliveryNote.id == note_id, DeliveryNote.deleted_at.is_(None))
    )
    note = result.scalar_one_or_none()
    if note is None:
        return fail("Delivery note not found", 404)
    data = _dn_row(note)
    items_result = await db.execute(
        select(DeliveryNoteItem).where(
            DeliveryNoteItem.delivery_note_id == note_id, DeliveryNoteItem.deleted_at.is_(None)
        )
    )
    data["items"] = [{
        "id": i.id, "delivery_note_id": i.delivery_note_id,
        "product_id": i.product_id, "quantity": i.quantity,
    } for i in items_result.scalars().all()]
    return ok(data)


@dn_router.post("", status_code=201)
async def create_delivery_note(body: DeliveryNoteCreate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    data = body.model_dump()
    if not data.get("note_no"):
        data["note_no"] = await _next_seq(db, "DN", DeliveryNote, "note_no")
    for date_field in ("delivery_date", "signed_at"):
        if data.get(date_field):
            data[date_field] = _parse_date(data[date_field])
    note = DeliveryNote(**data)
    db.add(note)
    await db.flush()
    return ok({"id": note.id, "note_no": note.note_no})


@dn_router.put("/{note_id}")
async def update_delivery_note(note_id: int, body: DeliveryNoteUpdate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(DeliveryNote).where(DeliveryNote.id == note_id, DeliveryNote.deleted_at.is_(None))
    )
    note = result.scalar_one_or_none()
    if note is None:
        return fail("Delivery note not found", 404)
    data = body.model_dump(exclude_unset=True)
    for date_field in ("delivery_date", "signed_at"):
        if data.get(date_field):
            data[date_field] = _parse_date(data[date_field])
    for key, val in data.items():
        setattr(note, key, val)
    await db.flush()
    return ok({"id": note.id})


@dn_router.delete("/{note_id}")
async def delete_delivery_note(note_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(DeliveryNote).where(DeliveryNote.id == note_id, DeliveryNote.deleted_at.is_(None))
    )
    note = result.scalar_one_or_none()
    if note is None:
        return fail("Delivery note not found", 404)
    note.deleted_at = datetime.now(timezone.utc)
    await db.flush()
    return ok(msg="deleted")


# --- Delivery Note Items ---

@dn_router.get("/{note_id}/items")
async def list_delivery_note_items(note_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    rows = (await db.execute(
        select(DeliveryNoteItem).where(
            DeliveryNoteItem.delivery_note_id == note_id, DeliveryNoteItem.deleted_at.is_(None)
        )
    )).scalars().all()
    return ok([{
        "id": i.id, "delivery_note_id": i.delivery_note_id,
        "product_id": i.product_id, "quantity": i.quantity,
    } for i in rows])


@dn_router.post("/{note_id}/items", status_code=201)
async def create_delivery_note_item(note_id: int, body: DeliveryNoteItemCreate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    item = DeliveryNoteItem(delivery_note_id=note_id, **body.model_dump())
    db.add(item)
    await db.flush()
    return ok({"id": item.id, "product_id": item.product_id})


@dn_router.put("/{note_id}/items/{item_id}")
async def update_delivery_note_item(note_id: int, item_id: int, body: DeliveryNoteItemUpdate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(DeliveryNoteItem).where(
            DeliveryNoteItem.id == item_id, DeliveryNoteItem.delivery_note_id == note_id,
            DeliveryNoteItem.deleted_at.is_(None),
        )
    )
    item = result.scalar_one_or_none()
    if item is None:
        return fail("Delivery note item not found", 404)
    for key, val in body.model_dump(exclude_unset=True).items():
        setattr(item, key, val)
    await db.flush()
    return ok({"id": item.id})


@dn_router.delete("/{note_id}/items/{item_id}")
async def delete_delivery_note_item(note_id: int, item_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(DeliveryNoteItem).where(
            DeliveryNoteItem.id == item_id, DeliveryNoteItem.delivery_note_id == note_id,
            DeliveryNoteItem.deleted_at.is_(None),
        )
    )
    item = result.scalar_one_or_none()
    if item is None:
        return fail("Delivery note item not found", 404)
    item.deleted_at = datetime.now(timezone.utc)
    await db.flush()
    return ok(msg="deleted")


# --- Sales Funnel ---

@opp_router.get("/funnel")
async def get_sales_funnel(
    customer_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    base = select(
        Opportunity.stage,
        func.count(Opportunity.id),
        func.coalesce(func.sum(Opportunity.amount), 0),
    ).where(Opportunity.deleted_at.is_(None))
    if customer_id:
        base = base.where(Opportunity.customer_id == customer_id)
    base = base.group_by(Opportunity.stage)

    rows = (await db.execute(base)).all()
    stage_order = {"lead": 0, "qualified": 1, "proposal": 2, "negotiation": 3, "won": 4, "lost": 5}
    funnel = []
    for row in rows:
        funnel.append({"stage": row[0], "count": row[1], "amount": float(row[2])})
    funnel.sort(key=lambda x: stage_order.get(x["stage"], 99))
    return ok(funnel)


# --- Sales Stats ---

stats_router = APIRouter(prefix="/sales", tags=["sales-stats"])


@stats_router.get("/stats/summary")
async def get_sales_summary(db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    order_stats = (await db.execute(
        select(
            func.count(SalesOrder.id),
            func.coalesce(func.sum(SalesOrder.total_amount), 0),
            func.coalesce(func.avg(SalesOrder.total_amount), 0),
        ).where(SalesOrder.deleted_at.is_(None))
    )).first()

    active_opps = (await db.execute(
        select(func.count(Opportunity.id)).where(
            Opportunity.deleted_at.is_(None),
            Opportunity.stage.notin_(["won", "lost"]),
        )
    )).scalar() or 0

    return ok({
        "total_orders": order_stats[0] or 0,
        "total_amount": float(order_stats[1]),
        "avg_amount": round(float(order_stats[2]), 2),
        "active_opportunities": active_opps,
    })


@stats_router.get("/stats/trend")
async def get_sales_trend(
    period: str = Query("monthly", pattern="^(monthly|quarterly)$"),
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    trunc = "month" if period == "monthly" else "quarter"
    fmt = "YYYY-MM" if period == "monthly" else "YYYY-Q"

    conditions = [SalesOrder.deleted_at.is_(None)]
    if start_date:
        conditions.append(SalesOrder.created_at >= _parse_date(start_date))
    if end_date:
        conditions.append(SalesOrder.created_at <= _parse_date(end_date))

    rows = (await db.execute(
        select(
            func.date_trunc(trunc, SalesOrder.created_at).label("period"),
            func.count(SalesOrder.id),
            func.coalesce(func.sum(SalesOrder.total_amount), 0),
        ).where(*conditions).group_by(text("period")).order_by(text("period"))
    )).all()

    return ok([{
        "period": str(row[0])[:7] if period == "monthly" else f"{row[0].year}-Q{(row[0].month - 1) // 3 + 1}",
        "order_count": row[1],
        "total_amount": float(row[2]),
    } for row in rows])


@stats_router.get("/stats/stage-distribution")
async def get_stage_distribution(db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    rows = (await db.execute(
        select(
            Opportunity.stage,
            func.count(Opportunity.id),
        ).where(Opportunity.deleted_at.is_(None)).group_by(Opportunity.stage)
    )).all()
    total = sum(r[1] for r in rows) or 1
    return ok([{
        "stage": r[0],
        "count": r[1],
        "percentage": round(r[1] / total * 100, 1),
    } for r in rows])


# --- Flow Conversion ---

@quo_router.post("/{quo_id}/convert-to-order")
async def convert_quotation_to_order(quo_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    quo_result = await db.execute(
        select(Quotation).where(Quotation.id == quo_id, Quotation.deleted_at.is_(None))
    )
    quo = quo_result.scalar_one_or_none()
    if quo is None:
        return fail("Quotation not found", 404)

    items_result = await db.execute(
        select(QuotationItem).where(QuotationItem.quotation_id == quo_id, QuotationItem.deleted_at.is_(None))
    )
    quo_items = items_result.scalars().all()

    order_no = await _next_seq(db, "SO", SalesOrder, "order_no")
    order = SalesOrder(
        order_no=order_no,
        customer_id=quo.customer_id,
        status="pending",
        total_amount=quo.total_amount,
        notes=f"Converted from quotation {quo.quotation_no}",
    )
    db.add(order)
    await db.flush()

    for qi in quo_items:
        item = SalesOrderItem(
            order_id=order.id,
            product_id=qi.product_id,
            quantity=qi.quantity,
            unit_price=qi.unit_price,
            amount=qi.amount,
        )
        db.add(item)

    quo.status = "approved"
    await db.flush()
    return ok({"id": order.id, "document_no": order.order_no, "msg": f"报价单 {quo.quotation_no} 已转为销售订单"})


@so_router.post("/{order_id}/convert-to-delivery")
async def convert_order_to_delivery(order_id: int, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    order_result = await db.execute(
        select(SalesOrder).where(SalesOrder.id == order_id, SalesOrder.deleted_at.is_(None))
    )
    order = order_result.scalar_one_or_none()
    if order is None:
        return fail("Sales order not found", 404)

    items_result = await db.execute(
        select(SalesOrderItem).where(SalesOrderItem.order_id == order_id, SalesOrderItem.deleted_at.is_(None))
    )
    order_items = items_result.scalars().all()

    note_no = await _next_seq(db, "DN", DeliveryNote, "note_no")
    note = DeliveryNote(
        note_no=note_no,
        sales_order_id=order.id,
        customer_id=order.customer_id,
        status="pending",
        notes=f"Converted from order {order.order_no}",
    )
    db.add(note)
    await db.flush()

    for oi in order_items:
        item = DeliveryNoteItem(
            delivery_note_id=note.id,
            product_id=oi.product_id,
            quantity=oi.quantity,
        )
        db.add(item)

    order.status = "confirmed"
    await db.flush()
    return ok({"id": note.id, "document_no": note.note_no, "msg": f"销售订单 {order.order_no} 已转为送货单"})


# --- Batch Operations ---

@opp_router.post("/batch-delete")
async def batch_delete_opportunities(body: BatchDeleteRequest, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    result = await db.execute(
        select(Opportunity).where(Opportunity.id.in_(body.ids), Opportunity.deleted_at.is_(None))
    )
    rows = result.scalars().all()
    for row in rows:
        row.deleted_at = now
    await db.flush()
    return ok({"deleted": len(rows)})


@opp_router.post("/batch-update")
async def batch_update_opportunities(body: OpportunityBatchUpdate, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    result = await db.execute(
        select(Opportunity).where(Opportunity.id.in_(body.ids), Opportunity.deleted_at.is_(None))
    )
    rows = result.scalars().all()
    for row in rows:
        if body.stage is not None:
            row.stage = body.stage
        if body.probability is not None:
            row.probability = body.probability
    await db.flush()
    return ok({"updated": len(rows)})


@quo_router.post("/batch-delete")
async def batch_delete_quotations(body: BatchDeleteRequest, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    result = await db.execute(
        select(Quotation).where(Quotation.id.in_(body.ids), Quotation.deleted_at.is_(None))
    )
    rows = result.scalars().all()
    for row in rows:
        row.deleted_at = now
    await db.flush()
    return ok({"deleted": len(rows)})


@so_router.post("/batch-delete")
async def batch_delete_sales_orders(body: BatchDeleteRequest, db: AsyncSession = Depends(get_db), _user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    result = await db.execute(
        select(SalesOrder).where(SalesOrder.id.in_(body.ids), SalesOrder.deleted_at.is_(None))
    )
    rows = result.scalars().all()
    for row in rows:
        row.deleted_at = now
    await db.flush()
    return ok({"deleted": len(rows)})


# --- Excel Import/Export ---

QUO_EXPORT_HEADERS = ["报价单号", "客户ID", "状态", "总金额", "有效期至", "备注", "创建时间"]
SO_EXPORT_HEADERS = ["订单号", "客户ID", "状态", "总金额", "交货日期", "备注", "创建时间"]


def _export_excel(headers: list[str], rows: list[list], filename: str) -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@quo_router.get("/export")
async def export_quotations(
    customer_id: int | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    base = select(Quotation).where(Quotation.deleted_at.is_(None))
    if customer_id:
        base = base.where(Quotation.customer_id == customer_id)
    if status:
        base = base.where(Quotation.status == status)
    rows = (await db.execute(base.order_by(Quotation.id.desc()))).scalars().all()
    data = [[
        q.quotation_no or "",
        q.customer_id,
        q.status,
        float(q.total_amount),
        str(q.valid_until)[:10] if q.valid_until else "",
        q.notes or "",
        str(q.created_at)[:19],
    ] for q in rows]
    return _export_excel(QUO_EXPORT_HEADERS, data, "quotations.xlsx")


@quo_router.post("/import")
async def import_quotations(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        data = {
            "customer_id": int(row[1]) if row[1] else 0,
            "status": str(row[2]) if row[2] else "draft",
            "total_amount": float(row[3]) if row[3] else 0,
            "notes": str(row[5]) if len(row) > 5 and row[5] else None,
        }
        if len(row) > 4 and row[4]:
            data["valid_until"] = str(row[4])
        data["quotation_no"] = await _next_seq(db, "QO", Quotation, "quotation_no")
        quo = Quotation(**data)
        db.add(quo)
        imported += 1
    await db.flush()
    return ok({"imported": imported})


@so_router.get("/export")
async def export_sales_orders(
    customer_id: int | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    base = select(SalesOrder).where(SalesOrder.deleted_at.is_(None))
    if customer_id:
        base = base.where(SalesOrder.customer_id == customer_id)
    if status:
        base = base.where(SalesOrder.status == status)
    rows = (await db.execute(base.order_by(SalesOrder.id.desc()))).scalars().all()
    data = [[
        s.order_no or "",
        s.customer_id,
        s.status,
        float(s.total_amount),
        str(s.delivery_date)[:10] if s.delivery_date else "",
        s.notes or "",
        str(s.created_at)[:19],
    ] for s in rows]
    return _export_excel(SO_EXPORT_HEADERS, data, "sales_orders.xlsx")


@so_router.post("/import")
async def import_sales_orders(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _user: dict = Depends(get_current_user),
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        data = {
            "customer_id": int(row[1]) if row[1] else 0,
            "status": str(row[2]) if row[2] else "pending",
            "total_amount": float(row[3]) if row[3] else 0,
            "notes": str(row[5]) if len(row) > 5 and row[5] else None,
        }
        if len(row) > 4 and row[4]:
            data["delivery_date"] = str(row[4])
        data["order_no"] = await _next_seq(db, "SO", SalesOrder, "order_no")
        order = SalesOrder(**data)
        db.add(order)
        imported += 1
    await db.flush()
    return ok({"imported": imported})
